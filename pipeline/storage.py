"""Excel persistence adapter with Power BI formatting and atomic replacement."""
from __future__ import annotations
import re
import shutil
import tempfile
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from .exceptions import MasterDataError, MissingColumnError

def inspect_source(path: Path, sheet: str) -> list[str]:
    """Read source headers without allowing pandas to hide duplicates."""
    if not path.exists():
        raise FileNotFoundError(f"Source file does not exist: {path}")
    if path.suffix.lower() == ".xls":
        # xlrd 2.x deliberately supports the legacy binary .xls format only.
        import xlrd
        workbook = xlrd.open_workbook(path, on_demand=True)
        if sheet not in workbook.sheet_names():
            workbook.release_resources()
            raise MissingColumnError(f"Required worksheet '{sheet}' is absent from {path.name}")
        worksheet = workbook.sheet_by_name(sheet)
        headers = worksheet.row_values(0) if worksheet.nrows else []
        workbook.release_resources()
        return headers
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        workbook.close()
        raise MissingColumnError(f"Required worksheet '{sheet}' is absent from {path.name}")
    headers = [cell.value for cell in next(workbook[sheet].iter_rows(min_row=1, max_row=1))]
    workbook.close()
    return headers

def load_source_file(path: Path, sheet: str) -> pd.DataFrame:
    """Load a validated source and retain the physical Excel row number."""
    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    df = pd.read_excel(path, sheet_name=sheet, dtype=object, engine=engine)
    df["Source_Row_Number"] = range(2, len(df) + 2)
    return df

def empty_table(columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=columns)

def load_master_table(path: Path, columns: list[str]) -> pd.DataFrame:
    """Load an existing master or return its configured empty shape."""
    # Prefer the pipeline's current .xlsx output, but migrate a same-named
    # legacy .xls master automatically when that is all the user has supplied.
    actual_path = path if path.exists() else path.with_suffix(".xls")
    if not actual_path.exists():
        return empty_table(columns)
    try:
        # Object dtype is essential for natural keys such as "000001"; pandas'
        # type inference would otherwise turn them into integer 1 on reload.
        engine = "xlrd" if actual_path.suffix.lower() == ".xls" else "openpyxl"
        frame = pd.read_excel(actual_path, sheet_name="Data", dtype=object, engine=engine)
    except Exception as exc:
        raise MasterDataError(f"Cannot read master table {actual_path}: {exc}") from exc
    missing = set(columns) - set(frame.columns)
    if missing:
        raise MasterDataError(f"Master {actual_path.name} is missing columns: {sorted(missing)}")
    return frame[columns]

def write_excel_table(df: pd.DataFrame, path: Path, table_name: str) -> None:
    """Write a clean Excel table with filters, frozen header and useful widths."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
    workbook = load_workbook(path)
    ws = workbook["Data"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if len(df):
        safe_name = re.sub(r"[^A-Za-z0-9_]", "_", table_name)
        table = Table(displayName=safe_name[:255], ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    for cells in ws.columns:
        width = min(50, max(10, max(len(str(c.value)) if c.value is not None else 0 for c in cells) + 2))
        ws.column_dimensions[cells[0].column_letter].width = width
    workbook.save(path)

def backup_master_files(master_folder: Path, archive_folder: Path, stamp: str, names: list[str]) -> Path | None:
    """Copy existing masters immediately before a changing commit."""
    existing = []
    for name in names:
        existing.extend(path for path in (master_folder / f"{name}.xlsx", master_folder / f"{name}.xls") if path.exists())
    if not existing:
        return None
    destination = archive_folder / stamp
    destination.mkdir(parents=True, exist_ok=False)
    for path in existing:
        shutil.copy2(path, destination / path.name)
    return destination

def save_master_tables(tables: dict[str, pd.DataFrame], master_folder: Path) -> None:
    """Stage and verify every workbook before atomically replacing master files."""
    master_folder.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(dir=master_folder.parent) as temp_dir:
        temp = Path(temp_dir)
        for name, frame in tables.items():
            target = temp / f"{name}.xlsx"
            write_excel_table(frame, target, name)
            # Verify with the storage engine itself. This avoids applying pandas
            # inference during transaction validation and checks physical rows.
            workbook = load_workbook(target, read_only=True, data_only=True)
            sheet = workbook["Data"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            row_count = max(0, sheet.max_row - 1)
            workbook.close()
            if row_count != len(frame) or headers != list(frame.columns):
                raise MasterDataError(f"Staged workbook verification failed: {name}")
        for name in tables:
            (temp / f"{name}.xlsx").replace(master_folder / f"{name}.xlsx")

def save_rejected_rows(frame: pd.DataFrame, folder: Path, stamp: str) -> Path | None:
    if frame.empty:
        return None
    path = folder / f"rejected_{stamp}.xlsx"
    write_excel_table(frame, path, f"Rejected_{stamp}")
    return path
